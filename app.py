import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
from database import init_db, SessionLocal, SalesRecord, UploadLog, BacklogSnapshot, BudgetRecord
from data_processor import process_sales_upload, process_annual_upload, process_budget_upload
from sqlalchemy import func

# Initialize DB on first load
init_db()

st.set_page_config(page_title="Sales Dashboard", layout="wide", page_icon="📈")

# ── PASSWORD GATE ──────────────────────────────────────────────────────────────
# Set APP_PASSWORD in .streamlit/secrets.toml before deploying.
# Local fallback is "admin" — change it!
_PASSWORD = st.secrets.get("APP_PASSWORD", "admin")

if not st.session_state.get("authenticated"):
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] > .main { display:flex; align-items:center; justify-content:center; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("<br><br>", unsafe_allow_html=True)
    col_l, col_c, col_r = st.columns([2, 1, 2])
    with col_c:
        st.image("https://img.icons8.com/color/96/000000/combo-chart--v1.png", width=72)
        st.markdown("### Sales Dashboard")
        pwd = st.text_input("Password", type="password", key="_login_pwd")
        if st.button("Sign in", use_container_width=True):
            if pwd == _PASSWORD:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    st.stop()
# ──────────────────────────────────────────────────────────────────────────────

# --- CUSTOM CSS ---
st.markdown("""
<style>
    .metric-card {
        background-color: #f8f9fa;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        text-align: center;
        border-left: 5px solid #1f77b4;
    }
    .metric-value {
        font-size: 2.2rem;
        font-weight: 700;
        color: #2c3e50;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #7f8c8d;
        text-transform: uppercase;
        font-weight: 600;
        letter-spacing: 1px;
    }
    .stApp {
        background-color: #ffffff;
    }
    .file-slot {
        border-radius: 8px;
        padding: 8px 12px 4px 12px;
        margin-bottom: 6px;
    }
    .file-slot-pending {
        background-color: #f0f4f8;
        border-left: 4px solid #b0bec5;
    }
    .file-slot-done {
        background-color: #e8f5e9;
        border-left: 4px solid #43a047;
    }
    .file-slot-label {
        font-size: 0.78rem;
        font-weight: 700;
        letter-spacing: 0.5px;
        margin-bottom: 2px;
    }
    .file-slot-label-pending { color: #546e7a; }
    .file-slot-label-done { color: #2e7d32; }
    .file-status {
        font-size: 0.72rem;
        margin-top: -4px;
        margin-bottom: 4px;
    }
    .file-status-pending { color: #90a4ae; }
    .file-status-done { color: #43a047; font-weight: 600; }

    /* ── Widget labels (text above selectboxes, radios, etc.) ── */
    [data-testid="stWidgetLabel"] p,
    [data-testid="stWidgetLabel"] {
        font-size: 1.05rem !important;
        font-weight: 600 !important;
    }

    /* ── Selectbox selected value and dropdown options ── */
    [data-baseweb="select"] span,
    [data-baseweb="select"] div,
    [data-baseweb="menu"] li {
        font-size: 1.05rem !important;
    }

    /* ── Radio button option labels ── */
    [data-testid="stRadio"] label p,
    [data-testid="stRadio"] label span {
        font-size: 1.05rem !important;
    }

    /* ── General markdown text in main area ── */
    [data-testid="stMarkdownContainer"] p {
        font-size: 1.05rem !important;
    }

    /* ── Tab labels ── */
    [data-testid="stTabs"] button p {
        font-size: 1.05rem !important;
        font-weight: 600 !important;
    }

    /* ── Dataframe / ag-grid table font size ── */
    .stDataFrame .ag-cell,
    .stDataFrame .ag-header-cell-text,
    .stDataFrame .ag-cell-value {
        font-size: 1rem !important;
        line-height: 1.5 !important;
    }
    .stDataFrame .ag-row {
        min-height: 36px !important;
    }
    .stDataFrame .ag-header-row {
        min-height: 40px !important;
    }
</style>
""", unsafe_allow_html=True)

import os

# --- SIDEBAR CONTROL PANEL ---
with st.sidebar:
    st.image("https://cdn-icons-png.flaticon.com/512/3050/3050525.png", width=60)
    st.title("Data Management")
    st.markdown("Upload the 5 monthly `.xlsx` exports below. The system automatically categorizes and joins them.")
    
    st.subheader("1. Monthly Data Upload")
    upload_year = st.selectbox("Assign to Year", [2024, 2025, 2026, 2027], index=2)
    upload_month = st.selectbox("Assign to Month", range(1, 13), format_func=lambda x: ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][x-1])

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    FILE_SLOTS = [
        ("f_job_summary", "Job Summary of Billings & Cost",  "Job Summary of Billings and Cost MMDDYY.xlsx"),
        ("f_gl",          "General Ledger",                  "General Ledger MMDDYY.xlsx"),
        ("f_sales_person","Sales Analysis by Sales Person",  "Sales Analysis by Sales Person MMDDYY.xlsx"),
        ("f_inv_item",    "Sales Analysis by Inventory Item","Sales Analysis by Inventory Item MMDDYY.xlsx"),
        ("f_job_cost",    "Job Cost Query",                  "Job Cost Query MMDDYY.xlsx"),
    ]

    uploaded = {}
    for key, label, hint in FILE_SLOTS:
        f = st.session_state.get(key)
        is_done = f is not None
        slot_cls = "file-slot-done" if is_done else "file-slot-pending"
        lbl_cls  = "file-slot-label-done" if is_done else "file-slot-label-pending"
        status_cls = "file-status-done" if is_done else "file-status-pending"
        status_txt = "✔ Ready" if is_done else "Awaiting upload…"

        st.markdown(f"""
        <div class="file-slot {slot_cls}">
            <div class="file-slot-label {lbl_cls}">{label}</div>
            <div class="file-status {status_cls}">{status_txt}</div>
        </div>
        """, unsafe_allow_html=True)
        uploaded[key] = st.file_uploader(hint, type=["xlsx"], key=key, label_visibility="collapsed")

    all_files = [v for v in uploaded.values() if v is not None]
    ready = len(all_files) == 5

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    if not ready:
        st.caption(f"{len(all_files)}/5 files uploaded")

    if ready:
        if st.button("Process Monthly Data", type="primary", use_container_width=True):
            with st.spinner("Joining files & updating records..."):
                success, msg = process_sales_upload(all_files, upload_year, upload_month)
                if success:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)
                    
    st.markdown("---")
    st.subheader("2. Annual / Bulk Upload")
    st.caption("Upload one set of full-year files. The system will split them into monthly records automatically.")

    annual_year = st.selectbox("Year", [2023, 2024, 2025, 2026], index=2, key="annual_year")

    ANNUAL_SLOTS = [
        ("af_job_summary", "Job Summary of Billings & Cost",  "Job Summary (full year).xlsx"),
        ("af_gl",          "General Ledger",                  "General Ledger (full year).xlsx"),
        ("af_sales_person","Sales Analysis by Sales Person",  "Sales Analysis by Sales Person (full year).xlsx"),
        ("af_inv_item",    "Sales Analysis by Inventory Item","Sales Analysis by Inventory Item (full year).xlsx"),
        ("af_job_cost",    "Job Cost Query",                  "Job Cost Query (full year).xlsx"),
    ]

    annual_uploaded = {}
    for key, label, hint in ANNUAL_SLOTS:
        f = st.session_state.get(key)
        is_done = f is not None
        slot_cls   = "file-slot-done"    if is_done else "file-slot-pending"
        lbl_cls    = "file-slot-label-done"  if is_done else "file-slot-label-pending"
        status_cls = "file-status-done"  if is_done else "file-status-pending"
        status_txt = "✔ Ready"           if is_done else "Awaiting upload…"
        st.markdown(f"""
        <div class="file-slot {slot_cls}">
            <div class="file-slot-label {lbl_cls}">{label}</div>
            <div class="file-status {status_cls}">{status_txt}</div>
        </div>
        """, unsafe_allow_html=True)
        annual_uploaded[key] = st.file_uploader(hint, type=["xlsx"], key=key, label_visibility="collapsed")

    annual_files = [v for v in annual_uploaded.values() if v is not None]
    annual_ready = len(annual_files) == 5

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    if not annual_ready:
        st.caption(f"{len(annual_files)}/5 files uploaded")

    if annual_ready:
        if st.button("Process Full Year", type="primary", use_container_width=True, key="btn_annual"):
            with st.spinner(f"Splitting {annual_year} data by month and loading…"):
                success, msg = process_annual_upload(annual_files, annual_year)
                if success:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    st.markdown("---")
    st.subheader("3. Budget Upload")

    budget_year = st.selectbox("Budget Year", [2024, 2025, 2026, 2027], index=2, key="budget_year")

    bf = st.session_state.get("budget_file_key")
    bf_done = bf is not None
    st.markdown(f"""
    <div class="file-slot {'file-slot-done' if bf_done else 'file-slot-pending'}">
        <div class="file-slot-label {'file-slot-label-done' if bf_done else 'file-slot-label-pending'}">Annual Budget File</div>
        <div class="file-status {'file-status-done' if bf_done else 'file-status-pending'}">{'✔ Ready' if bf_done else 'Awaiting upload…'}</div>
    </div>
    """, unsafe_allow_html=True)
    budget_file = st.file_uploader("budget_YYYY.xlsx", type=["xlsx"], key="budget_file_key", label_visibility="collapsed")

    if budget_file is not None:
        if st.button("Load Budget", type="primary", use_container_width=True, key="btn_budget"):
            with st.spinner(f"Loading {budget_year} budget…"):
                success, msg = process_budget_upload(budget_file, budget_year)
                if success:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    st.markdown("---")
    st.subheader("4. Loaded Periods")

    session = SessionLocal()
    loaded = (
        session.query(SalesRecord.year, SalesRecord.month,
                      func.count(SalesRecord.id).label("jobs"),
                      func.sum(SalesRecord.invoiced).label("revenue"))
        .group_by(SalesRecord.year, SalesRecord.month)
        .order_by(SalesRecord.year, SalesRecord.month)
        .all()
    )
    last_upload = session.query(UploadLog).order_by(UploadLog.id.desc()).first()
    session.close()

    month_labels = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    if loaded:
        for row in loaded:
            col_a, col_b = st.columns([3, 1])
            with col_a:
                st.caption(f"{month_labels[row.month-1]} {row.year} — {row.jobs} jobs · ${row.revenue:,.0f}")
            with col_b:
                if st.button("🗑", key=f"del_{row.year}_{row.month}", help=f"Delete {month_labels[row.month-1]} {row.year}"):
                    session2 = SessionLocal()
                    session2.query(SalesRecord).filter(SalesRecord.year == row.year, SalesRecord.month == row.month).delete()
                    session2.query(BacklogSnapshot).filter(BacklogSnapshot.snapshot_year == row.year, BacklogSnapshot.snapshot_month == row.month).delete()
                    session2.commit()
                    session2.close()
                    st.rerun()
    else:
        st.caption("No data uploaded yet.")

    if last_upload:
        st.caption(f"Last upload: {last_upload.upload_timestamp.strftime('%Y-%m-%d %H:%M')} UTC")

# --- MAIN DASHBOARD ---
logo_path = "logo.png" if os.path.exists("logo.png") else ("logo.jpg" if os.path.exists("logo.jpg") else None)
if logo_path:
    # Use columns to elegantly center the extra-wide logo banner
    _, col_logo, _ = st.columns([1, 2, 1])
    with col_logo:
        st.image(logo_path, use_container_width=True)
        st.markdown("<br>", unsafe_allow_html=True)

st.title("Sales Performance Dashboard")
st.markdown("Analyze monthly revenue, gross profit, and budget comparisons.")

# Fetch Data
session = SessionLocal()
sales_df   = pd.read_sql(session.query(SalesRecord).statement,    session.bind)
backlog_df = pd.read_sql(session.query(BacklogSnapshot).statement, session.bind)
budget_df  = pd.read_sql(session.query(BudgetRecord).statement,   session.bind)
session.close()

def to_excel(df, currency_cols=None, pct_cols=None):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    currency_cols = set(currency_cols or [])
    pct_cols      = set(pct_cols      or [])

    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='004987', end_color='004987', fill_type='solid')
    alt_fill = PatternFill(start_color='EEF4FB', end_color='EEF4FB', fill_type='solid')

    # ── Headers ──────────────────────────────────────────────────────────────
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, (col, val) in enumerate(zip(df.columns, row), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if fill:
                c.fill = fill
            if col in currency_cols:
                c.number_format = '$#,##0'
                c.alignment     = Alignment(horizontal='right')
            elif col in pct_cols:
                c.number_format = '0.0"%"'
                c.alignment     = Alignment(horizontal='right')
            else:
                c.alignment = Alignment(horizontal='left')

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci, col in enumerate(df.columns, 1):
        letter = get_column_letter(ci)
        if col in currency_cols:
            max_val = pd.to_numeric(df[col], errors='coerce').abs().max()
            val_w   = len(f'${max_val:,.0f}') if pd.notna(max_val) else 12
            width   = max(len(col) + 2, val_w + 2)
        elif col in pct_cols:
            width = max(len(col) + 2, 8)
        else:
            max_len = df[col].astype(str).str.len().max() if len(df) > 0 else 10
            width   = max(len(col) + 2, min(int(max_len) + 2 if pd.notna(max_len) else 10, 45))
        ws.column_dimensions[letter].width = width

    wb.save(buf)
    return buf.getvalue()

def sortable_table(df, currency_cols=None, pct_cols=None, height=420):
    """Render a DataFrame as a sortable HTML table inside a components iframe."""
    currency_cols = set(currency_cols or [])
    pct_cols      = set(pct_cols      or [])
    tid = "t" + str(abs(hash(tuple(df.columns))))[:8]

    th_style  = "background:#f0f4f8;padding:11px 14px;font-size:15px;font-weight:700;border-bottom:2px solid #d0d7de;cursor:pointer;user-select:none;white-space:nowrap;position:sticky;top:0;z-index:1;"
    th_r      = th_style + "text-align:right;"
    td_style  = "padding:10px 14px;font-size:15px;border-bottom:1px solid #eef0f2;"
    td_r      = td_style + "text-align:right;"

    headers = ""
    for i, col in enumerate(df.columns):
        align_style = th_r if (col in currency_cols or col in pct_cols) else th_style
        headers += (f'<th style="{align_style}" onclick="sortBy({i})">'
                    f'{col} <span id="{tid}-icon-{i}" style="color:#aaa;font-size:11px;">&#8597;</span></th>')

    rows = ""
    for _, row in df.iterrows():
        rows += "<tr>"
        for col in df.columns:
            val = row[col]
            if col in currency_cols:
                try:    display = f"${float(val):,.0f}"
                except: display = str(val)
                rows += f'<td style="{td_r}" data-val="{val}">{display}</td>'
            elif col in pct_cols:
                try:    display = f"{float(val):.1f}%"
                except: display = str(val)
                rows += f'<td style="{td_r}" data-val="{val}">{display}</td>'
            else:
                rows += f'<td style="{td_style}" data-val="{str(val)}">{val}</td>'
        rows += "</tr>"

    html = f"""
    <style>
      body {{ margin:0; font-family:sans-serif; }}
      #wrap {{ overflow-x:auto; max-height:{height-10}px; overflow-y:auto; }}
      table {{ width:100%; border-collapse:collapse; }}
      tr:hover td {{ background:#f4f7fb; }}
    </style>
    <div id="wrap">
      <table id="{tid}">
        <thead><tr>{headers}</tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>
    <script>
      var _dir = {{}};
      function sortBy(col) {{
        var tbl = document.getElementById('{tid}');
        var tbody = tbl.querySelector('tbody');
        var rows = Array.from(tbody.querySelectorAll('tr'));
        _dir[col] = (_dir[col] === 'asc') ? 'desc' : 'asc';
        var asc = _dir[col] === 'asc';
        rows.sort(function(a, b) {{
          var av = a.cells[col].getAttribute('data-val');
          var bv = b.cells[col].getAttribute('data-val');
          var an = parseFloat(av), bn = parseFloat(bv);
          if (!isNaN(an) && !isNaN(bn)) return asc ? an - bn : bn - an;
          return asc ? av.localeCompare(bv) : bv.localeCompare(av);
        }});
        rows.forEach(function(r) {{ tbody.appendChild(r); }});
        // update icons
        tbl.querySelectorAll('th span[id]').forEach(function(s, i) {{
          s.innerHTML = i === col ? (asc ? '&#9650;' : '&#9660;') : '&#8597;';
          s.style.color = i === col ? '#004987' : '#aaa';
        }});
      }}
    </script>"""
    components.html(html, height=height, scrolling=False)

if sales_df.empty:
    st.info("Welcome! Please upload the 5 `.xlsx` Sales files in the sidebar to get started.")
else:
    tab1, tab2 = st.tabs(["Sales Dashboard", "Performance & Backlog"])

    month_names = {1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun',
                   7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'}

    # Shared HTML table cell styles
    _th   = "padding:11px 14px;font-size:1.05rem;font-weight:700;background:#f0f4f8;border-bottom:2px solid #d0d7de;text-align:left;white-space:nowrap;position:sticky;top:0;z-index:1;"
    _th_r = "padding:11px 14px;font-size:1.05rem;font-weight:700;background:#f0f4f8;border-bottom:2px solid #d0d7de;text-align:right;white-space:nowrap;position:sticky;top:0;z-index:1;"
    _td   = "padding:10px 14px;font-size:1rem;border-bottom:1px solid #eef0f2;"
    _tdr  = "padding:10px 14px;font-size:1rem;border-bottom:1px solid #eef0f2;text-align:right;"

    # =========================================================
    # TAB 1 — SALES DASHBOARD
    # =========================================================
    with tab1:
        st.markdown("### Filters")

        # ── Available date range in the data ──────────────────────────────────
        all_years_t1 = sorted(sales_df['year'].unique())
        max_year_t1  = max(all_years_t1)
        min_year_t1  = min(all_years_t1)
        last_data_month_t1 = int(sales_df[sales_df['year'] == max_year_t1]['month'].max())

        tf_col, sp_col = st.columns([3, 1])
        with tf_col:
            time_frame = st.radio("Time Frame", ["YTD", "Last Year", "Custom Range"],
                                  horizontal=True, key="t1_tf")

        with sp_col:
            salespeople = sorted(sales_df['salesperson'].dropna().unique().tolist())
            selected_sp = st.multiselect("Select Salesperson", salespeople,
                                         placeholder="All Salespeople", key="t1_sp")

        # ── Resolve start/end year+month from selected time frame ─────────────
        if time_frame == "YTD":
            start_year, start_month = max_year_t1, 1
            end_year,   end_month   = max_year_t1, last_data_month_t1
            period_label = f"YTD {max_year_t1}"

        elif time_frame == "Last Year":
            ly = max_year_t1 - 1
            start_year, start_month = ly, 1
            end_year,   end_month   = ly, 12
            period_label = str(ly)

        else:  # Custom Range
            cr1, cr2 = st.columns(2)
            with cr1:
                st.markdown("**Start**")
                start_year  = st.selectbox("Year ",  all_years_t1,
                                           index=0, key="t1_sy")
                avail_start = sorted(sales_df[sales_df['year'] == start_year]['month'].unique())
                start_month = st.selectbox("Month ", avail_start,
                                           format_func=lambda m: month_names[m],
                                           index=0, key="t1_sm")
            with cr2:
                st.markdown("**End**")
                end_year    = st.selectbox("Year  ", all_years_t1,
                                           index=len(all_years_t1) - 1, key="t1_ey")
                avail_end   = sorted(sales_df[sales_df['year'] == end_year]['month'].unique())
                end_month   = st.selectbox("Month  ", avail_end,
                                           format_func=lambda m: month_names[m],
                                           index=len(avail_end) - 1, key="t1_em")
            period_label = (f"{month_names[start_month]} {start_year} – "
                            f"{month_names[end_month]} {end_year}")

        # ── Filter dataframe by year+month range ──────────────────────────────
        ym           = sales_df['year'] * 100 + sales_df['month']
        start_ym     = start_year * 100 + start_month
        end_ym       = end_year   * 100 + end_month
        filtered_df  = sales_df[(ym >= start_ym) & (ym <= end_ym)]
        if selected_sp:
            filtered_df = filtered_df[filtered_df['salesperson'].isin(selected_sp)]
        safe_label = period_label.replace(" ", "_").replace("–", "-")

        st.markdown("---")

        total_rev    = filtered_df['invoiced'].sum()
        total_cost   = filtered_df['cost'].sum()
        total_profit = filtered_df['gross_profit'].sum()
        margin       = (total_profit / total_rev * 100) if total_rev > 0 else 0

        total_labor_rev  = filtered_df['labor_income'].sum()
        total_labor_cost = filtered_df['labor_cost'].sum()
        labor_margin     = ((total_labor_rev - total_labor_cost) / total_labor_rev * 100) if total_labor_rev > 0 else 0
        total_rental     = filtered_df['rental_income'].sum()

        kpi1, kpi2, kpi3 = st.columns(3)
        with kpi1:
            st.markdown(f"""
            <div class="kpi-card" style="border-left: 5px solid #004987;">
                <p>TOTAL REVENUE ({period_label})</p>
                <h2>${total_rev:,.0f}</h2>
            </div>
            """, unsafe_allow_html=True)
        with kpi2:
            st.markdown(f"""
            <div class="kpi-card" style="border-left: 5px solid #CF2E2E;">
                <p>TOTAL COST</p>
                <h2>${total_cost:,.0f}</h2>
            </div>
            """, unsafe_allow_html=True)
        with kpi3:
            st.markdown(f"""
            <div class="kpi-card" style="border-left: 5px solid #00D084;">
                <p>GROSS PROFIT MARGIN</p>
                <h2>{margin:.1f}%</h2>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        kpi4, kpi5, kpi6 = st.columns(3)
        with kpi4:
            st.markdown(f"""
            <div class="kpi-card" style="border-left: 5px solid #1565C0;">
                <p>LABOR REVENUE</p>
                <h2>${total_labor_rev:,.0f}</h2>
            </div>
            """, unsafe_allow_html=True)
        with kpi5:
            st.markdown(f"""
            <div class="kpi-card" style="border-left: 5px solid #6A1B9A;">
                <p>LABOR MARGIN</p>
                <h2>{labor_margin:.1f}%</h2>
            </div>
            """, unsafe_allow_html=True)
        with kpi6:
            st.markdown(f"""
            <div class="kpi-card" style="border-left: 5px solid #E65100;">
                <p>RENTAL INCOME</p>
                <h2>${total_rental:,.0f}</h2>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Monthly revenue chart ─────────────────────────────────────────────
        # Build ordered (year, month) pairs covering the selected range
        chart_pairs = []
        for y in range(start_year, end_year + 1):
            m_lo = start_month if y == start_year else 1
            m_hi = end_month   if y == end_year   else 12
            for m in range(m_lo, m_hi + 1):
                chart_pairs.append((y, m))

        pairs_df   = pd.DataFrame(chart_pairs, columns=['year', 'month'])
        rev_grouped = filtered_df.groupby(['year', 'month'])['invoiced'].sum().reset_index()
        monthly_rev = pd.merge(pairs_df, rev_grouped, on=['year', 'month'], how='left').fillna(0)

        # Label: just month name for single-year ranges, "Mon YYYY" for multi-year
        if start_year == end_year:
            monthly_rev['Label'] = monthly_rev['month'].map(month_names)
        else:
            monthly_rev['Label'] = monthly_rev.apply(
                lambda r: f"{month_names[int(r['month'])]} {int(r['year'])}", axis=1)

        fig = px.line(monthly_rev, x='Label', y='invoiced', markers=True,
                      title=f"Monthly Revenue Trend ({period_label})",
                      labels={'invoiced': 'Revenue ($)', 'Label': ''})
        fig.update_traces(line_color='#004987', line_width=3, marker_size=8)
        fig.update_layout(plot_bgcolor='white', hovermode='x unified',
                          hoverlabel=dict(font_size=15, bgcolor='white'))
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")
        sp_label = ", ".join(selected_sp) if selected_sp else "All Salespeople"
        st.markdown(f"### Detail View: {sp_label}")

        c_head, c_dl = st.columns([6, 1])
        with c_head:
            st.markdown("#### Top 10 Customers")
        top_cust = filtered_df.groupby('customer')['invoiced'].sum().nlargest(10).reset_index()
        top_cust.columns = ['Customer Name', 'Total Sales']
        with c_dl:
            st.download_button("⬇", to_excel(top_cust, currency_cols=['Total Sales']),
                               file_name="top10_customers.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_top10")
        sortable_table(top_cust, currency_cols=['Total Sales'], height=480)

        st.markdown("---")
        aj_head, aj_dl = st.columns([6, 1])
        with aj_head:
            st.markdown("#### All Jobs (Detailed View)")
        num_cols = ['invoiced', 'rental_income', 'labor_income',
                    'cost', 'labor_cost', 'other_costs', 'gross_profit']
        jobs_num  = filtered_df.groupby('job_number')[num_cols].sum().reset_index()
        # For text fields, prefer the row where customer differs from description
        # (i.e. where a real customer name was populated from the Inventory Item file)
        def best_text(grp):
            real = grp[grp['customer'] != grp['description']]
            src  = real if not real.empty else grp
            return pd.Series({
                'customer':   src['customer'].iloc[0],
                'description': src['description'].iloc[0],
                'salesperson': src['salesperson'].iloc[0],
            })
        jobs_meta = filtered_df.groupby('job_number').apply(best_text).reset_index()
        all_jobs  = pd.merge(jobs_num, jobs_meta, on='job_number')
        all_jobs  = all_jobs.sort_values(by='invoiced', ascending=False)
        all_jobs['non_labor_rent'] = all_jobs['invoiced'] - all_jobs['labor_income'] - all_jobs['rental_income']
        all_jobs['gp_pct']         = all_jobs.apply(lambda r: r['gross_profit'] / r['invoiced'] * 100 if r['invoiced'] else 0, axis=1)
        all_jobs['labor_margin_pct'] = all_jobs.apply(lambda r: (r['labor_income'] - r['labor_cost']) / r['labor_income'] * 100 if r['labor_income'] else 0, axis=1)
        all_jobs = all_jobs[['job_number', 'customer', 'description', 'salesperson',
                              'invoiced', 'rental_income', 'labor_income', 'non_labor_rent',
                              'cost', 'labor_cost', 'other_costs',
                              'gross_profit', 'gp_pct', 'labor_margin_pct']]
        all_jobs.columns = ['Job', 'Customer', 'Description', 'Salesperson',
                            'Invoiced', 'Rental Income', 'Labor Income', 'Other Income',
                            'Total Cost', 'Labor Cost', 'Other Costs',
                            'Gross Profit', 'GP %', 'Labor Margin %']
        with aj_dl:
            st.download_button("⬇", to_excel(all_jobs,
                                              currency_cols=['Invoiced', 'Rental Income', 'Labor Income', 'Other Income',
                                                             'Total Cost', 'Labor Cost', 'Other Costs', 'Gross Profit'],
                                              pct_cols=['GP %', 'Labor Margin %']),
                               file_name=f"all_jobs_{safe_label}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_alljobs")
        jobs_height = min(55 + len(all_jobs) * 44, 1400)
        sortable_table(all_jobs,
                       currency_cols=['Invoiced', 'Rental Income', 'Labor Income', 'Other Income',
                                      'Total Cost', 'Labor Cost', 'Other Costs', 'Gross Profit'],
                       pct_cols=['GP %', 'Labor Margin %'],
                       height=jobs_height)

        st.markdown("---")
        st.download_button(
            label="Download Filtered Data as Excel",
            data=to_excel(filtered_df,
                          currency_cols=['invoiced', 'rental_income', 'labor_income',
                                         'cost', 'labor_cost', 'other_costs', 'gross_profit']),
            file_name=f"sales_export_{safe_label}.xlsx",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # =========================================================
    # TAB 2 — PERFORMANCE & BACKLOG
    # =========================================================
    with tab2:
        st.markdown("### Performance & Backlog")

        all_years = sorted(sales_df['year'].unique(), reverse=True)
        all_sp = ["All Salespeople"] + sorted(sales_df['salesperson'].dropna().unique().tolist())

        pb_col1, pb_col2 = st.columns([1, 3])
        with pb_col1:
            salespeople_t2 = sorted(sales_df['salesperson'].dropna().unique().tolist())
            pb_sp = st.multiselect("Select Salesperson", salespeople_t2,
                                   placeholder="All Salespeople", key="t2_sp")

        # Determine current and prior year from available data
        current_year = max(all_years)
        prior_year = current_year - 1

        # Filter sales data by salesperson
        def sp_filter(df, sp_list):
            if not sp_list:
                return df
            return df[df['salesperson'].isin(sp_list)]

        curr_df = sp_filter(sales_df[sales_df['year'] == current_year], pb_sp)
        prev_df = sp_filter(sales_df[sales_df['year'] == prior_year], pb_sp)

        # Monthly revenue arrays (indexed 1-12)
        def monthly_array(df):
            g = df.groupby('month')['invoiced'].sum()
            return [g.get(m, 0.0) for m in range(1, 13)]

        curr_monthly = monthly_array(curr_df)
        prev_monthly = monthly_array(prev_df)

        # ── Budget lookup ─────────────────────────────────────────────────────
        # For "All Salespeople" use the COMPANY row; for a named person use their
        # row if it exists, otherwise default to 0.
        def budget_array(sp_list, year):
            if budget_df.empty:
                return [0.0] * 12
            yr_bud = budget_df[budget_df['year'] == year]
            if not sp_list:
                rows = yr_bud[yr_bud['salesperson'] == 'COMPANY']
            else:
                names = [s.strip().upper() for s in sp_list]
                rows = yr_bud[yr_bud['salesperson'].isin(names)]
            if rows.empty:
                return [0.0] * 12
            g = rows.groupby('month')['amount'].sum()
            return [g.get(m, 0.0) for m in range(1, 13)]

        budget_monthly = budget_array(pb_sp, current_year)

        # --- MONTHLY COMPARISON TABLE (HTML for readable font + full control) ---
        st.markdown("---")
        month_cols = [month_names[m] for m in range(1, 13)]

        # Last month with actual current-year data (for chart cutoff)
        last_data_idx = max((i for i, v in enumerate(curr_monthly) if v > 0), default=-1)

        def fmt(v): return f"${v:,.0f}" if v else "—"

        rows_cfg = [
            (f'{prior_year} Revenue', prev_monthly,   '#f8f9fa', '#2c3e50'),
            (f'{current_year} Revenue', curr_monthly, '#eaf4fb', '#004987'),
            (f'{current_year} Budget',  budget_monthly,'#fff8f0','#e67e22'),
        ]
        header_cells = "".join(
            f"<th style='text-align:right;padding:10px 8px;font-size:1rem;'>{m}</th>"
            for m in month_cols
        )
        body_rows = ""
        for label, data, bg, color in rows_cfg:
            cells = "".join(
                f"<td style='text-align:right;padding:10px 8px;font-size:1.05rem;'>{fmt(v)}</td>"
                for v in data
            )
            total = fmt(sum(data))
            body_rows += f"""
            <tr style="background:{bg};border-bottom:1px solid #dee2e6;">
                <td style="font-weight:700;color:{color};padding:10px 14px;white-space:nowrap;font-size:1.05rem;">{label}</td>
                {cells}
                <td style="font-weight:700;text-align:right;padding:10px 8px;color:{color};font-size:1.05rem;">{total}</td>
            </tr>"""

        st.markdown(f"""
        <div style="overflow-x:auto;font-size:1.05rem;margin-bottom:12px;">
        <table style="width:100%;border-collapse:collapse;border:1px solid #dee2e6;border-radius:6px;overflow:hidden;">
          <thead>
            <tr style="background:#004987;color:white;">
              <th style="text-align:left;padding:11px 14px;font-size:1rem;">Metric</th>
              {header_cells}
              <th style="text-align:right;padding:11px 8px;font-size:1rem;">Total</th>
            </tr>
          </thead>
          <tbody>{body_rows}</tbody>
        </table>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")

        # --- HARD BACKLOG ---
        # Always use the most recent snapshot and recalculate hard backlog
        # live from revised_contract − billed_to_date so it reflects the
        # latest contract values, not a value frozen at upload time.
        if not backlog_df.empty:
            latest_year  = backlog_df['snapshot_year'].max()
            latest_month = backlog_df[backlog_df['snapshot_year'] == latest_year]['snapshot_month'].max()
            latest_snapshot = backlog_df[
                (backlog_df['snapshot_year'] == latest_year) &
                (backlog_df['snapshot_month'] == latest_month)
            ].copy()
            # Recalculate hard backlog from source columns
            latest_snapshot['hard_backlog'] = (
                latest_snapshot['revised_contract'] - latest_snapshot['billed_to_date']
            ).clip(lower=0)
            bl_filtered  = sp_filter(latest_snapshot, pb_sp)
            hard_backlog = bl_filtered[bl_filtered['is_open']]['hard_backlog'].sum()
        else:
            hard_backlog = 0.0

        # YTD figures (all months with data so far)
        ytd_actual = sum(curr_monthly)
        ytd_budget = sum(budget_monthly)
        prior_total = sum(prev_monthly)
        difference = ytd_budget - ytd_actual

        # --- WRAP-UP KPI BOX ---
        w1, w2, w3, w4 = st.columns(4)
        with w1:
            st.markdown(f"""
            <div class="metric-card" style="border-left: 5px solid #7f8c8d;">
                <div class="metric-label">{prior_year} Revenue</div>
                <div class="metric-value">${prior_total:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)
        with w2:
            st.markdown(f"""
            <div class="metric-card" style="border-left: 5px solid #004987;">
                <div class="metric-label">{current_year} YTD Actual</div>
                <div class="metric-value">${ytd_actual:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)
        with w3:
            st.markdown(f"""
            <div class="metric-card" style="border-left: 5px solid #e67e22;">
                <div class="metric-label">Hard Backlog</div>
                <div class="metric-value">${hard_backlog:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)
        with w4:
            diff_color = "#00D084" if difference >= 0 else "#CF2E2E"
            diff_label = "Budget vs Actual"
            st.markdown(f"""
            <div class="metric-card" style="border-left: 5px solid {diff_color};">
                <div class="metric-label">{diff_label}</div>
                <div class="metric-value">${difference:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # --- LINE CHART ---
        # Current year actual line stops at the last month with data
        curr_monthly_chart = [
            v if i <= last_data_idx else None
            for i, v in enumerate(curr_monthly)
        ]
        chart_df = pd.DataFrame({
            'Month': month_cols,
            f'{prior_year} Revenue': prev_monthly,
            f'{current_year} Revenue': curr_monthly_chart,
            f'{current_year} Budget': budget_monthly,
        })
        fig2 = px.line(
            chart_df.melt(id_vars='Month', var_name='Series', value_name='Amount'),
            x='Month', y='Amount', color='Series', markers=True,
            title=f"Revenue Comparison — {prior_year} vs {current_year}",
            labels={'Amount': 'Revenue ($)', 'Month': ''}
        )
        color_map = {
            f'{prior_year} Revenue': '#7f8c8d',
            f'{current_year} Revenue': '#004987',
            f'{current_year} Budget': '#e67e22',
        }
        for trace in fig2.data:
            trace.line.color = color_map.get(trace.name, '#000000')
            trace.line.width = 2
            if trace.name == f'{current_year} Budget':
                trace.line.dash = 'dash'
        fig2.update_layout(plot_bgcolor='white', hovermode='x unified', legend_title_text='',
                           hoverlabel=dict(font_size=15, bgcolor='white'))
        fig2.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        st.plotly_chart(fig2, use_container_width=True)

        # --- BACKLOG DETAIL TABLE ---
        if not backlog_df.empty and hard_backlog > 0:
            st.markdown("---")
            bl_head, bl_dl = st.columns([6, 1])
            with bl_head:
                st.markdown("#### Open Jobs — Backlog Detail")
            bl_open = bl_filtered[bl_filtered['is_open'] & (bl_filtered['hard_backlog'] > 0)].copy()
            # Ensure hard_backlog reflects latest contract values
            bl_open['hard_backlog'] = (bl_open['revised_contract'] - bl_open['billed_to_date']).clip(lower=0)
            bl_open = bl_open[bl_open['hard_backlog'] > 0]
            bl_open = bl_open.sort_values('hard_backlog', ascending=False)
            bl_display = bl_open[['job_number', 'description', 'project_manager', 'salesperson',
                                   'revised_contract', 'billed_to_date', 'hard_backlog']].copy()
            bl_display.columns = ['Job', 'Description', 'Project Manager', 'Salesperson',
                                   'Revised Contract', 'Billed to Date', 'Hard Backlog']
            with bl_dl:
                st.download_button("⬇", to_excel(bl_display,
                                                  currency_cols=['Revised Contract', 'Billed to Date', 'Hard Backlog']),
                                   file_name="backlog_detail.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_backlog")
            bl_height = min(55 + len(bl_display) * 44, 1400)
            sortable_table(bl_display,
                           currency_cols=['Revised Contract', 'Billed to Date', 'Hard Backlog'],
                           height=bl_height)
